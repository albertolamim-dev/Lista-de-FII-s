import yfinance as yf
import pandas as pd
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, NamedStyle

# Lista de ativos FII (com sufixo .SA)
ativos = [
    'BTLG11.SA', 'XPLG11.SA', 'BRCO11.SA', 'HGLG11.SA', 'VILG11.SA',
    'TRXF11.SA', 'GARE11.SA', 'HGRU11.SA', 'KNRI11.SA', 'XPML11.SA',
    'VISC11.SA', 'HGBS11.SA', 'HSML11.SA', 'JSRE11.SA', 'BRCR11.SA',
    'VINO11.SA', 'KNSC11.SA', 'RBRR11.SA', 'KNCR11.SA', 'RECR11.SA',
    'MXRF11.SA', 'BCRI11.SA', 'SNAG11.SA', 'BBGO11.SA', 'RZAG11.SA',
    'CPTR11.SA'
]

# Criar pasta relatorios se não existir
os.makedirs('relatorios', exist_ok=True)

dados = []
print(f"Iniciando extração de {len(ativos)} ativos...")

for ativo in ativos:
    try:
        ticker = yf.Ticker(ativo)
        info = ticker.info
        
        # Extração robusta
        price = info.get('currentPrice') or info.get('regularMarketPrice') or info.get('previousClose')
        dy_raw = info.get('dividendYield')
        pvp_raw = info.get('priceToBook')
        
        dados.append({
            'Ativo': ativo.replace('.SA', ''),
            'Cotação (R$)': round(price, 2) if price else None,
            'Dividend Yield': round(dy_raw, 4) if dy_raw else None,
            'P/VP': round(pvp_raw, 2) if pvp_raw else None
        })
        print(f"✅ {ativo} OK")
    except Exception as e:
        print(f"❌ Erro {ativo}: {e}")
        dados.append({'Ativo': ativo.replace('.SA', ''), 'Cotação (R$)': None, 'Dividend Yield': None, 'P/VP': None})

# Criar DataFrame e ordenar
df = pd.DataFrame(dados)
df = df.sort_values(by='Ativo').reset_index(drop=True)

# Configurar Excel
wb = Workbook()
ws = wb.active
ws.title = 'Monitoramento FIIs'

# Cabeçalhos
headers = list(df.columns)
for col_num, header in enumerate(headers, 1):
    ws.cell(row=1, column=col_num, value=header)

# Cores para formatação condicional
green_fill = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid') # Verde claro (< 1.00)
red_fill = PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')   # Vermelho claro (> 1.05)

# Estilo de porcentagem
percent_style = NamedStyle(name='percent', number_format='0.00%')
wb.add_named_style(percent_style)

# Escrever dados e aplicar estilos linha a linha
for r_idx, row in enumerate(df.itertuples(index=False), 2):
    ws.cell(row=r_idx, column=1, value=row[0]) # Ativo
    ws.cell(row=r_idx, column=2, value=row[1]) # Cotação
    
    # DY com porcentagem
    cell_dy = ws.cell(row=r_idx, column=3, value=row[2])
    if row[2] is not None:
        cell_dy.style = 'percent'
        
    # P/VP com cores
    cell_pvp = ws.cell(row=r_idx, column=4, value=row[3])
    if row[3] is not None:
        if row[3] < 1.00:
            cell_pvp.fill = green_fill
        elif row[3] > 1.05:
            cell_pvp.fill = red_fill

# Salvar arquivo com data
timestamp = datetime.now().strftime('%Y-%m-%d')
nome_arquivo = f'relatorios/Carteira_FIIs_{timestamp}.xlsx'
wb.save(nome_arquivo)
print(f"\n✅ Arquivo salvo com sucesso: {nome_arquivo}")
